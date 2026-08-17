from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
import shutil
from typing import Iterable

from openpyxl import Workbook, load_workbook


CATALOG_HEADERS = (
    "sku",
    "product_name",
    "unit_cost",
    "pack_size",
    "opening_stock",
    "current_stock",
)
TRANSACTION_HEADERS = (
    "transaction_id",
    "date",
    "sku",
    "direction",
    "package_count",
    "unit_cost",
    "line_total",
)
SAFE_ID = re.compile(r"[A-Za-z0-9][A-Za-z0-9._-]{0,63}\Z")
FORMULA_PREFIXES = ("=", "+", "-", "@")


class ReconciliationError(ValueError):
    """Raised when workbook data violates a reconciliation invariant."""


@dataclass(frozen=True)
class CatalogItem:
    row: int
    sku: str
    product_name: str
    unit_cost: Decimal
    pack_size: Decimal
    opening_stock: Decimal
    current_stock: Decimal


@dataclass(frozen=True)
class Transaction:
    transaction_id: str
    sku: str
    direction: str
    package_count: Decimal
    unit_cost: Decimal
    line_total: Decimal


@dataclass(frozen=True)
class ReconciledRow:
    item: CatalogItem
    inbound_units: Decimal
    outbound_units: Decimal
    expected_stock: Decimal


def reconcile(
    *,
    catalog_path: Path,
    transactions_path: Path,
    output_dir: Path,
    mode: str,
) -> dict[str, object]:
    """Validate, aggregate, report and optionally update an inventory catalog."""
    if mode not in {"check", "apply"}:
        raise ReconciliationError("mode must be 'check' or 'apply'")
    _require_xlsx_file(catalog_path, "catalog")
    _require_xlsx_file(transactions_path, "transactions")

    items = _read_catalog(catalog_path)
    transactions = _read_transactions(transactions_path, items)
    rows = _aggregate(items, transactions)

    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / f"reconciliation-{mode}.xlsx"
    _write_report(report_path, rows, mode)

    backup_path: Path | None = None
    if mode == "apply":
        backup_path = _apply_catalog(catalog_path, rows)

    return {
        "mode": mode,
        "catalog_items": len(items),
        "transactions": len(transactions),
        "report_path": str(report_path),
        "backup_path": str(backup_path) if backup_path else None,
    }


def _require_xlsx_file(path: Path, label: str) -> None:
    if path.suffix.lower() != ".xlsx" or not path.is_file():
        raise ReconciliationError(f"{label} must be an existing .xlsx file")


def _read_catalog(path: Path) -> dict[str, CatalogItem]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        sheet = workbook["Catalog"] if "Catalog" in workbook.sheetnames else workbook.active
        _require_headers(sheet, CATALOG_HEADERS)
        items: dict[str, CatalogItem] = {}
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(value in (None, "") for value in values[: len(CATALOG_HEADERS)]):
                continue
            sku = _safe_identifier(values[0], f"Catalog!A{row_number}")
            if sku in items:
                raise ReconciliationError(f"duplicate SKU: {sku}")
            name = _safe_text(values[1], f"Catalog!B{row_number}")
            item = CatalogItem(
                row=row_number,
                sku=sku,
                product_name=name,
                unit_cost=_positive_decimal(values[2], f"Catalog!C{row_number}"),
                pack_size=_positive_decimal(values[3], f"Catalog!D{row_number}"),
                opening_stock=_non_negative_decimal(values[4], f"Catalog!E{row_number}"),
                current_stock=_non_negative_decimal(values[5], f"Catalog!F{row_number}"),
            )
            items[sku] = item
        if not items:
            raise ReconciliationError("catalog contains no items")
        return items
    finally:
        workbook.close()


def _read_transactions(path: Path, items: dict[str, CatalogItem]) -> list[Transaction]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        sheet = workbook["Transactions"] if "Transactions" in workbook.sheetnames else workbook.active
        _require_headers(sheet, TRANSACTION_HEADERS)
        seen: set[str] = set()
        result: list[Transaction] = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(value in (None, "") for value in values[: len(TRANSACTION_HEADERS)]):
                continue
            transaction_id = _safe_identifier(values[0], f"Transactions!A{row_number}")
            if transaction_id in seen:
                raise ReconciliationError(f"duplicate transaction_id: {transaction_id}")
            seen.add(transaction_id)
            sku = _safe_identifier(values[2], f"Transactions!C{row_number}")
            if sku not in items:
                raise ReconciliationError(f"unknown SKU in transaction {transaction_id}: {sku}")
            direction = str(values[3] or "").strip().lower()
            if direction not in {"inbound", "outbound"}:
                raise ReconciliationError(f"invalid direction in transaction {transaction_id}")
            package_count = _positive_decimal(values[4], f"Transactions!E{row_number}")
            unit_cost = _positive_decimal(values[5], f"Transactions!F{row_number}")
            line_total = _positive_decimal(values[6], f"Transactions!G{row_number}")
            item = items[sku]
            units = package_count * item.pack_size
            if unit_cost != item.unit_cost:
                raise ReconciliationError(f"unit cost mismatch in transaction {transaction_id}")
            if line_total != units * unit_cost:
                raise ReconciliationError(f"line total mismatch in transaction {transaction_id}")
            result.append(Transaction(transaction_id, sku, direction, package_count, unit_cost, line_total))
        if not result:
            raise ReconciliationError("transactions workbook contains no rows")
        return result
    finally:
        workbook.close()


def _aggregate(
    items: dict[str, CatalogItem], transactions: Iterable[Transaction]
) -> list[ReconciledRow]:
    inbound = {sku: Decimal("0") for sku in items}
    outbound = {sku: Decimal("0") for sku in items}
    for transaction in transactions:
        units = transaction.package_count * items[transaction.sku].pack_size
        target = inbound if transaction.direction == "inbound" else outbound
        target[transaction.sku] += units

    rows: list[ReconciledRow] = []
    for sku, item in items.items():
        expected = item.opening_stock + inbound[sku] - outbound[sku]
        if expected < 0:
            raise ReconciliationError(f"negative resulting stock for SKU {sku}")
        rows.append(ReconciledRow(item, inbound[sku], outbound[sku], expected))
    return rows


def _write_report(path: Path, rows: list[ReconciledRow], mode: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reconciliation"
    sheet.append([
        "sku", "product_name", "opening_stock", "inbound_units",
        "outbound_units", "recorded_stock", "expected_stock", "variance", "mode",
    ])
    for row in rows:
        sheet.append([
            _excel_text(row.item.sku),
            _excel_text(row.item.product_name),
            float(row.item.opening_stock),
            float(row.inbound_units),
            float(row.outbound_units),
            float(row.item.current_stock),
            float(row.expected_stock),
            float(row.expected_stock - row.item.current_stock),
            mode,
        ])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _atomic_save(workbook, path)


def _apply_catalog(path: Path, rows: list[ReconciledRow]) -> Path:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    backup_path = path.with_name(f"{path.stem}.backup-{timestamp}{path.suffix}")
    shutil.copy2(path, backup_path)
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        sheet = workbook["Catalog"] if "Catalog" in workbook.sheetnames else workbook.active
        for row in rows:
            sheet.cell(row.item.row, 6).value = float(row.expected_stock)
        _atomic_save(workbook, path)
    except Exception:
        shutil.copy2(backup_path, path)
        raise
    return backup_path


def _atomic_save(workbook: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.stem}.tmp{path.suffix}")
    try:
        workbook.save(temporary)
        temporary.replace(path)
    finally:
        workbook.close()
        temporary.unlink(missing_ok=True)


def _require_headers(sheet: object, expected: tuple[str, ...]) -> None:
    values = tuple(str(cell.value or "").strip() for cell in sheet[1][: len(expected)])
    if values != expected:
        raise ReconciliationError(f"unexpected headers: expected {expected}, got {values}")


def _decimal(value: object, cell: str) -> Decimal:
    if isinstance(value, str) and value.lstrip().startswith("="):
        raise ReconciliationError(f"formula is not allowed in input cell {cell}")
    try:
        number = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError) as error:
        raise ReconciliationError(f"invalid number in {cell}") from error
    if not number.is_finite():
        raise ReconciliationError(f"non-finite number in {cell}")
    return number


def _positive_decimal(value: object, cell: str) -> Decimal:
    number = _decimal(value, cell)
    if number <= 0:
        raise ReconciliationError(f"value must be positive in {cell}")
    return number


def _non_negative_decimal(value: object, cell: str) -> Decimal:
    number = _decimal(value, cell)
    if number < 0:
        raise ReconciliationError(f"value must be non-negative in {cell}")
    return number


def _safe_identifier(value: object, cell: str) -> str:
    text = str(value or "").strip()
    if not SAFE_ID.fullmatch(text):
        raise ReconciliationError(f"invalid identifier in {cell}")
    return text


def _safe_text(value: object, cell: str) -> str:
    text = str(value or "").strip()
    if not text or len(text) > 120 or any(ord(char) < 32 for char in text):
        raise ReconciliationError(f"invalid text in {cell}")
    return text


def _excel_text(value: str) -> str:
    return f"'{value}" if value.startswith(FORMULA_PREFIXES) else value
