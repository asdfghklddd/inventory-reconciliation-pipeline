from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from inventory_reconciliation import ReconciliationError, reconcile


def build_catalog(path: Path, *, opening: int = 100, current: int = 100, name: str = "Demo item") -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(["sku", "product_name", "unit_cost", "pack_size", "opening_stock", "current_stock"])
    sheet.append(["SKU-001", name, 5, 2, opening, current])
    workbook.save(path)
    workbook.close()


def build_transactions(path: Path, *, sku: str = "SKU-001", direction: str = "outbound", packages: int = 3) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"
    sheet.append(["transaction_id", "date", "sku", "direction", "package_count", "unit_cost", "line_total"])
    sheet.append(["TX-001", "2026-08-01", sku, direction, packages, 5, packages * 2 * 5])
    workbook.save(path)
    workbook.close()


def test_check_mode_reports_without_mutating_catalog(tmp_path: Path) -> None:
    catalog = tmp_path / "catalog.xlsx"
    transactions = tmp_path / "transactions.xlsx"
    build_catalog(catalog)
    build_transactions(transactions)

    result = reconcile(catalog_path=catalog, transactions_path=transactions, output_dir=tmp_path / "out", mode="check")

    workbook = load_workbook(catalog, data_only=True)
    assert workbook["Catalog"]["F2"].value == 100
    workbook.close()
    assert Path(result["report_path"]).is_file()
    assert result["backup_path"] is None


def test_apply_mode_backs_up_and_updates_catalog(tmp_path: Path) -> None:
    catalog = tmp_path / "catalog.xlsx"
    transactions = tmp_path / "transactions.xlsx"
    build_catalog(catalog)
    build_transactions(transactions)

    result = reconcile(catalog_path=catalog, transactions_path=transactions, output_dir=tmp_path / "out", mode="apply")

    workbook = load_workbook(catalog, data_only=True)
    assert workbook["Catalog"]["F2"].value == 94
    workbook.close()
    assert Path(result["backup_path"]).is_file()


def test_unknown_sku_fails_closed(tmp_path: Path) -> None:
    catalog = tmp_path / "catalog.xlsx"
    transactions = tmp_path / "transactions.xlsx"
    build_catalog(catalog)
    build_transactions(transactions, sku="SKU-404")

    with pytest.raises(ReconciliationError, match="unknown SKU"):
        reconcile(catalog_path=catalog, transactions_path=transactions, output_dir=tmp_path / "out", mode="check")


def test_negative_resulting_stock_is_rejected(tmp_path: Path) -> None:
    catalog = tmp_path / "catalog.xlsx"
    transactions = tmp_path / "transactions.xlsx"
    build_catalog(catalog, opening=2, current=2)
    build_transactions(transactions, packages=2)

    with pytest.raises(ReconciliationError, match="negative resulting stock"):
        reconcile(catalog_path=catalog, transactions_path=transactions, output_dir=tmp_path / "out", mode="check")


def test_formula_input_is_rejected(tmp_path: Path) -> None:
    catalog = tmp_path / "catalog.xlsx"
    transactions = tmp_path / "transactions.xlsx"
    build_catalog(catalog)
    build_transactions(transactions)
    workbook = load_workbook(transactions)
    workbook["Transactions"]["E2"] = "=1+1"
    workbook.save(transactions)
    workbook.close()

    with pytest.raises(ReconciliationError, match="formula is not allowed"):
        reconcile(catalog_path=catalog, transactions_path=transactions, output_dir=tmp_path / "out", mode="check")
